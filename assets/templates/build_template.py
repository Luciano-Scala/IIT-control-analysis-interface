"""
assets/templates/build_template.py
"""

from __future__ import annotations

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_ORIGINAL_PATH = "/mnt/user-data/uploads/Preinforme_indentación.xlsm"
_OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "FM-086_preinforme_indentacion.xlsx")

_THIN = Side(style="thin", color="000000")
_MEDIUM = Side(style="medium", color="000000")
_FONT_HEADER = Font(name="Arial", size=10, bold=True)
_FONT_DATA = Font(name="Arial", size=10, bold=False)
_FONT_FORMULA = Font(name="Arial", size=9, bold=True, italic=True, color="1F4E78")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_CENTER_NOWRAP = Alignment(horizontal="center", vertical="center")
_FILL_HEADER = PatternFill("solid", fgColor="FFD9D9D9")

# Columnas de la tabla de resultados rediseñada
_COL_OT = "A"
_COL_NRO = "B"
_COL_SIGMA_Y = "C"
_COL_ERR_SIGMA_Y = "D"
_COL_UTS = "E"
_COL_ERR_UTS = "F"
_COL_CARGA_DUREZA = "G"
_COL_DUREZA = "H"
_COL_OBS_INICIO = "I"
_COL_OBS_FIN = "L"

_FILA_HEADER = 43
_FILA_DATA_INICIO = 44
_FILA_DATA_FIN = 52
_FILA_TOTAL = 53
_FILA_DESV = 54


def _border_celda(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN) -> Border:
    return Border(left=left, right=right, top=top, bottom=bottom)


def build() -> str:
    wb = openpyxl.load_workbook(_ORIGINAL_PATH, keep_vba=False, data_only=False)
    ws = wb["FM-086"]

    _quitar_logo(ws)
    _actualizar_titulo(ws)
    _reconstruir_tabla_resultados(ws)
    _reconstruir_memoria_analitica(ws)

    wb.save(_OUTPUT_PATH)
    return _OUTPUT_PATH


def _quitar_logo(ws) -> None:
    """El área A1:C3 queda vacía — sin logo de LABTESA, disponible para
    que se pegue un logo distinto por fuera de la app."""
    ws._images = []


def _actualizar_titulo(ws) -> None:
    """"MODELO DE INFORME N° 1" era una etiqueta de plantilla de
    ejemplo; se reemplaza por un campo real de N° de informe."""
    ws["D1"] = "INFORME N°:"


def _reconstruir_tabla_resultados(ws) -> None:
    # Deshacer los merges de 3 columnas anchas (uno por fila de header+datos)
    for fila in range(_FILA_HEADER, _FILA_DATA_FIN + 1):
        for rango in (f"A{fila}:B{fila}", f"D{fila}:E{fila}", f"F{fila}:H{fila}", f"I{fila}:L{fila}"):
            try:
                ws.unmerge_cells(rango)
            except KeyError:
                pass
    # TOTAL/DESVIACIÓN conservan el merge A:B para la etiqueta (igual
    # que el original), solo se deshacen los merges de D:E/F:H/I:L.
    for fila in (_FILA_TOTAL, _FILA_DESV):
        for rango in (f"D{fila}:E{fila}", f"F{fila}:H{fila}", f"I{fila}:L{fila}"):
            try:
                ws.unmerge_cells(rango)
            except KeyError:
                pass

    # Anchos de columna: A-E se dejan EXACTAMENTE como en el original,
    # porque la sección "1.1- VERIFICACIONES" (más arriba en la misma
    # hoja) reutiliza esas mismas columnas para texto libre que se
    # derrama sobre celdas vacías vecinas — cambiarlas ahí rompe esa
    # sección. F/G/H no se usan en ninguna otra parte de la hoja, así
    # que sí se pueden ajustar con libertad (F era un simple espaciador
    # de 3 caracteres dentro del merge F:H original, inutilizable como
    # columna de datos propia).
    ws.column_dimensions["F"].width = 9.5

    # --- Encabezado (fila 43) ---
    headers = {
        _COL_OT: "OT",
        _COL_NRO: "N°",
        _COL_SIGMA_Y: "σy\n[MPa]",
        _COL_ERR_SIGMA_Y: "Δσy\n[MPa]",
        _COL_UTS: "UTS\n[MPa]",
        _COL_ERR_UTS: "ΔUTS\n[MPa]",
        _COL_CARGA_DUREZA: "Carga\ndureza [N]",
        _COL_DUREZA: "Dureza\n[HB]",
    }
    for col, texto in headers.items():
        c = ws[f"{col}{_FILA_HEADER}"]
        c.value = texto
        c.font = _FONT_HEADER
        c.alignment = _ALIGN_CENTER
        c.fill = _FILL_HEADER
        c.border = _border_celda(top=_MEDIUM, bottom=_MEDIUM, left=_MEDIUM if col == _COL_OT else _THIN,
                                   right=_MEDIUM if col == _COL_DUREZA else _THIN)

    ws.merge_cells(f"{_COL_OBS_INICIO}{_FILA_HEADER}:{_COL_OBS_FIN}{_FILA_HEADER}")
    obs_header = ws[f"{_COL_OBS_INICIO}{_FILA_HEADER}"]
    obs_header.value = "Observaciones"
    obs_header.font = _FONT_HEADER
    obs_header.alignment = _ALIGN_CENTER
    obs_header.fill = _FILL_HEADER
    obs_header.border = _border_celda(top=_MEDIUM, bottom=_MEDIUM, right=_MEDIUM)

    # --- Filas de datos (44-52): solo formato + bordes; los valores los
    # escribe utils/excel_exporter.py. Estas son las celdas de "input". ---
    for fila in range(_FILA_DATA_INICIO, _FILA_DATA_FIN + 1):
        for col in (_COL_OT, _COL_NRO, _COL_SIGMA_Y, _COL_ERR_SIGMA_Y, _COL_UTS, _COL_ERR_UTS,
                    _COL_CARGA_DUREZA, _COL_DUREZA):
            c = ws[f"{col}{fila}"]
            c.font = _FONT_DATA
            c.alignment = _ALIGN_CENTER_NOWRAP
            c.border = _border_celda(left=_MEDIUM if col == _COL_OT else _THIN,
                                       right=_MEDIUM if col == _COL_DUREZA else _THIN)
            if col in (_COL_SIGMA_Y, _COL_ERR_SIGMA_Y, _COL_UTS, _COL_ERR_UTS):
                c.number_format = "0"
            elif col == _COL_CARGA_DUREZA:
                c.number_format = "0"
            elif col == _COL_DUREZA:
                c.number_format = "0.0"

        ws.merge_cells(f"{_COL_OBS_INICIO}{fila}:{_COL_OBS_FIN}{fila}")
        obs = ws[f"{_COL_OBS_INICIO}{fila}"]
        obs.font = _FONT_DATA
        obs.alignment = Alignment(horizontal="left", vertical="center")
        obs.border = _border_celda(right=_MEDIUM)

    # --- TOTAL (fila 53): cuenta de indentaciones y de improntas ---
    rango_sigma = f"{_COL_SIGMA_Y}{_FILA_DATA_INICIO}:{_COL_SIGMA_Y}{_FILA_DATA_FIN}"
    rango_dureza = f"{_COL_DUREZA}{_FILA_DATA_INICIO}:{_COL_DUREZA}{_FILA_DATA_FIN}"

    ws[f"{_COL_SIGMA_Y}{_FILA_TOTAL}"] = f"=COUNT({rango_sigma})"
    ws[f"{_COL_DUREZA}{_FILA_TOTAL}"] = f"=COUNT({rango_dureza})"
    for col in (_COL_SIGMA_Y, _COL_DUREZA):
        c = ws[f"{col}{_FILA_TOTAL}"]
        c.font = _FONT_FORMULA
        c.alignment = _ALIGN_CENTER_NOWRAP
        c.number_format = "0"

    # --- DESVIACIÓN (fila 54): coeficiente de variación % de cada magnitud ---
    rango_uts = f"{_COL_UTS}{_FILA_DATA_INICIO}:{_COL_UTS}{_FILA_DATA_FIN}"

    ws[f"{_COL_SIGMA_Y}{_FILA_DESV}"] = f'=IFERROR(STDEV({rango_sigma})/AVERAGE({rango_sigma})*100,"-")'
    ws[f"{_COL_UTS}{_FILA_DESV}"] = f'=IFERROR(STDEV({rango_uts})/AVERAGE({rango_uts})*100,"-")'
    ws[f"{_COL_DUREZA}{_FILA_DESV}"] = f'=IFERROR(STDEV({rango_dureza})/AVERAGE({rango_dureza})*100,"-")'
    for col in (_COL_SIGMA_Y, _COL_UTS, _COL_DUREZA):
        c = ws[f"{col}{_FILA_DESV}"]
        c.font = _FONT_FORMULA
        c.alignment = _ALIGN_CENTER_NOWRAP
        c.number_format = '0.0"%"'


def _reconstruir_memoria_analitica(ws) -> None:
    """"MEMORIA ANALÍTICA" pasa a mostrar fórmulas que promedian
    directamente la tabla de resultados, en vez de texto fijo escrito
    por Python."""
    rango_sigma = f"{_COL_SIGMA_Y}{_FILA_DATA_INICIO}:{_COL_SIGMA_Y}{_FILA_DATA_FIN}"
    rango_err_sigma = f"{_COL_ERR_SIGMA_Y}{_FILA_DATA_INICIO}:{_COL_ERR_SIGMA_Y}{_FILA_DATA_FIN}"
    rango_dureza = f"{_COL_DUREZA}{_FILA_DATA_INICIO}:{_COL_DUREZA}{_FILA_DATA_FIN}"

    ws["I11"] = (
        f'=IF(COUNT({rango_sigma})=0,"(sin datos)",'
        f'TEXT(AVERAGE({rango_sigma}),"0")&"  ±  "&TEXT(AVERAGE({rango_err_sigma}),"0")&"  MPa")'
    )
    ws["I11"].font = Font(name="Arial", size=10, bold=True)
    ws["I11"].alignment = _ALIGN_CENTER_NOWRAP

    ws["I12"] = (
        f'=IF(COUNT({rango_dureza})=0,"DUREZA HB: (sin datos)",'
        f'"DUREZA HB: "&TEXT(AVERAGE({rango_dureza}),"0.0"))'
    )
    ws["I12"].font = Font(name="Arial", size=10, bold=False)
    ws["I12"].alignment = Alignment(horizontal="left", vertical="center")

    # El bloque "ARCHIVO GUARDADO EN" (I13:L15) no se usa — un informe
    # agregado puede combinar indentaciones de varios CSV distintos, no
    # hay un único archivo de origen que mostrar (ver decisión en
    # utils/excel_exporter.py). Se deja como nota libre en su lugar.
    ws.unmerge_cells("I14:L15")
    ws.merge_cells("I13:L15")
    obs_gen = ws["I13"]
    obs_gen.value = "Observaciones generales:"
    obs_gen.font = _FONT_HEADER
    obs_gen.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


if __name__ == "__main__":
    path = build()
    print(f"Template generado en: {path}")