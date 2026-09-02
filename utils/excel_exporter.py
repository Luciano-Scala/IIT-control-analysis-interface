"""
utils/excel_exporter.py
=========================
Exportación del Pre-Informe de Indentación al formulario **FM-086**, a
partir del template en
``assets/templates/FM-086_preinforme_indentacion.xlsx``.
"""

from __future__ import annotations

import io
import os
from dataclasses import dataclass, field
from typing import List, Optional

import openpyxl
from openpyxl.drawing.image import Image as XLImage

_TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "assets", "templates", "FM-086_preinforme_indentacion.xlsx",
)

_SHEET_NAME = "FM-086"
_TABLA_FILA_INICIO = 44
_TABLA_FILA_FIN = 52  # 9 filas disponibles en el template

# Columnas de la tabla "1.4- RESULTADOS OBTENIDOS" (ver
# assets/templates/build_template.py para el layout completo)
_COL_NRO = "B"
_COL_SIGMA_Y = "C"
_COL_ERR_SIGMA_Y = "D"
_COL_UTS = "E"
_COL_ERR_UTS = "F"
_COL_CARGA_DUREZA = "G"
_COL_DUREZA = "H"

_GRAFICO_ANCHOR = "A21"
# Tamaño en píxeles del área en blanco "1.3- GRÁFICO DE INDENTACIÓN"
# (A21:L40), calculado a partir de los anchos de columna y altos de
# fila reales del template.
_GRAFICO_ANCHO_PX = 847
_GRAFICO_ALTO_PX = 400


@dataclass
class IndentacionRow:
    """Una fila de indentación para la tabla 1.4 (columnas σy / UTS)."""
    sigma_y_mpa: float
    err_sigma_y_mpa: float
    uts_mpa: float
    err_uts_mpa: float


@dataclass
class ImprontaRow:
    """Una impronta Brinell para la tabla 1.4 (columnas Carga / Dureza).
    ``dureza_hb`` es el promedio de los 3 diámetros medidos en esa
    impronta."""
    carga_N: float
    dureza_hb: float


@dataclass
class DatosInforme:
    """Todo lo que necesita el exportador, ya extraído de la UI (sin
    objetos Qt)."""
    cliente: str = ""
    solicitud: str = ""
    fecha: str = ""
    muestra: str = ""
    lugar: str = ""
    coordenadas: str = ""
    temperatura: str = ""
    comentario: str = ""
    indentaciones: List[IndentacionRow] = field(default_factory=list)
    improntas: List[ImprontaRow] = field(default_factory=list)
    grafico_png_bytes: Optional[bytes] = None


def exportar_informe(datos: DatosInforme, output_path: str) -> str:
    """Genera el informe FM-086 completo a partir del template y lo
    guarda en ``output_path``. Devuelve la ruta final.

    Las fórmulas del template (TOTAL, DESVIACIÓN, MEMORIA ANALÍTICA) se
    recalculan solas la primera vez que Excel/LibreOffice abre el
    archivo — este módulo no necesita (ni debe) tocarlas."""
    wb = openpyxl.load_workbook(_TEMPLATE_PATH)
    ws = wb[_SHEET_NAME]

    _llenar_encabezado(ws, datos)
    _llenar_tabla_resultados(ws, datos)
    if datos.grafico_png_bytes:
        _insertar_grafico(ws, datos.grafico_png_bytes)

    wb.save(output_path)
    return output_path


def _llenar_encabezado(ws, datos: DatosInforme) -> None:
    """Completa los campos "label: valor en la misma celda", igual al
    convenio ya usado en el template (ver CLIENTE/SOLICITUD/FECHA)."""
    if datos.cliente:
        ws["I1"] = f"CLIENTE : {datos.cliente}"
    if datos.solicitud:
        ws["I2"] = f"SOLICITUD N°: {datos.solicitud}"
    if datos.fecha:
        ws["I3"] = f"FECHA : {datos.fecha}"
    if datos.muestra:
        ws["D5"] = f"MUESTRA:  {datos.muestra}"
    if datos.lugar:
        ws["A13"] = f"LUGAR:  {datos.lugar}"
    if datos.coordenadas:
        ws["A11"] = datos.coordenadas
    if datos.temperatura:
        ws["D12"] = f"TEMPERATURA: {datos.temperatura}°C"
    if datos.comentario:
        ws["A57"] = datos.comentario


def _llenar_tabla_resultados(ws, datos: DatosInforme) -> None:
    """Llena "1.4- RESULTADOS OBTENIDOS" con números reales (no texto):
    las columnas de indentación (σy/Δσy/UTS/ΔUTS) y de dureza
    (Carga/Dureza) se completan de forma independiente por fila — si
    una lista es más corta que la otra, esas celdas quedan en blanco.
    TOTAL/DESVIACIÓN/MEMORIA ANALÍTICA son fórmulas del template que
    leen estas mismas celdas; no se tocan acá."""
    n_filas_disponibles = _TABLA_FILA_FIN - _TABLA_FILA_INICIO + 1
    n_filas = min(max(len(datos.indentaciones), len(datos.improntas)), n_filas_disponibles)

    for i in range(n_filas):
        fila = _TABLA_FILA_INICIO + i
        ws[f"{_COL_NRO}{fila}"] = i + 1

        if i < len(datos.indentaciones):
            ind = datos.indentaciones[i]
            ws[f"{_COL_SIGMA_Y}{fila}"] = round(ind.sigma_y_mpa)
            ws[f"{_COL_ERR_SIGMA_Y}{fila}"] = round(ind.err_sigma_y_mpa)
            ws[f"{_COL_UTS}{fila}"] = round(ind.uts_mpa)
            ws[f"{_COL_ERR_UTS}{fila}"] = round(ind.err_uts_mpa)

        if i < len(datos.improntas):
            imp = datos.improntas[i]
            ws[f"{_COL_CARGA_DUREZA}{fila}"] = round(imp.carga_N)
            ws[f"{_COL_DUREZA}{fila}"] = round(imp.dureza_hb, 1)


def _insertar_grafico(ws, png_bytes: bytes) -> None:
    img = XLImage(io.BytesIO(png_bytes))
    img.width = _GRAFICO_ANCHO_PX
    img.height = _GRAFICO_ALTO_PX
    ws.add_image(img, _GRAFICO_ANCHOR)